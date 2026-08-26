import requests
import pandas as pd
import time
import re
from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

START_SEASON = 2021
END_SEASON = 2025

OUTPUT_FILE = (
    f"MotoGP_Historical_Race_"
    f"{START_SEASON}_{END_SEASON}.xlsx"
)

BASE_URL = "https://api.motogp.pulselive.com/motogp/v1"

HEADERS = {
    "User-Agent": "Mozilla/5.0",
    "Accept": "application/json",
    "Origin": "https://www.motogp.com",
    "Referer": "https://www.motogp.com/"
}

REQUEST_DELAY = 0.5


# ============================================================
# REQUEST FUNCTION
# ============================================================

def get_json(url, params=None, retries=3):

    for attempt in range(1, retries + 1):

        try:

            response = requests.get(
                url,
                params=params,
                headers=HEADERS,
                timeout=30
            )

            response.raise_for_status()

            return response.json()

        except requests.exceptions.RequestException as e:

            print(
                f"      Request error "
                f"({attempt}/{retries}): {e}"
            )

            if attempt < retries:
                time.sleep(2)

    return None


# ============================================================
# GET SEASON ID
# ============================================================

def get_season_id(year):

    url = f"{BASE_URL}/results/seasons"

    data = get_json(url)

    if not data:
        return None

    # API bisa mengembalikan list langsung
    if isinstance(data, list):

        seasons = data

    # atau dictionary
    elif isinstance(data, dict):

        seasons = (
            data.get("seasons")
            or data.get("content")
            or data.get("results")
            or []
        )

    else:

        seasons = []

    for season in seasons:

        try:

            season_year = int(
                season.get("year")
            )

        except (
            ValueError,
            TypeError
        ):

            continue

        if season_year == year:

            return (
                season.get("id")
                or season.get("uuid")
            )

    return None


# ============================================================
# GET MOTOGP CATEGORY
# ============================================================

def get_motogp_category_id(season_uuid):

    url = f"{BASE_URL}/results/categories"

    params = {
        "seasonUuid": season_uuid
    }

    data = get_json(
        url,
        params
    )

    if not data:
        return None

    if isinstance(data, list):

        categories = data

    elif isinstance(data, dict):

        categories = (
            data.get("categories")
            or data.get("content")
            or data.get("results")
            or []
        )

    else:

        categories = []

    for category in categories:

        name = str(
            category.get("name", "")
        ).lower()

        if name == "motogp":

            return (
                category.get("id")
                or category.get("uuid")
            )

    # fallback
    for category in categories:

        name = str(
            category.get("name", "")
        ).lower()

        if "motogp" in name:

            return (
                category.get("id")
                or category.get("uuid")
            )

    return None


# ============================================================
# GET EVENTS
# ============================================================

def get_events(season_uuid):

    url = f"{BASE_URL}/results/events"

    params = {
        "seasonUuid": season_uuid,
        "isFinished": "true"
    }

    data = get_json(
        url,
        params
    )

    if not data:
        return []

    if isinstance(data, list):

        return data

    if isinstance(data, dict):

        return (
            data.get("events")
            or data.get("content")
            or data.get("results")
            or []
        )

    return []


# ============================================================
# GET SESSIONS
# ============================================================

def get_sessions(
    event_uuid,
    category_uuid
):

    url = f"{BASE_URL}/results/sessions"

    params = {
        "eventUuid": event_uuid,
        "categoryUuid": category_uuid
    }

    data = get_json(
        url,
        params
    )

    if not data:
        return []

    if isinstance(data, list):

        return data

    if isinstance(data, dict):

        return (
            data.get("sessions")
            or data.get("content")
            or data.get("results")
            or []
        )

    return []


# ============================================================
# FIND MAIN RACE
# ============================================================

def find_race_session(sessions):

    """
    Hanya mencari Main Race.

    Sprint tidak diambil.
    """

    for session in sessions:

        session_type = str(
            session.get("type", "")
        ).upper()

        session_name = str(
            session.get("name", "")
        ).upper()

        # Main Race biasanya type = RAC
        if session_type == "RAC":

            return session

        # fallback
        if (
            session_name == "RACE"
            or session_name == "MOTOGP™ RACE"
        ):

            return session

    return None


# ============================================================
# GET CLASSIFICATION
# ============================================================

def get_classification(
    session_id,
    season
):

    url = (
        f"{BASE_URL}/results/session/"
        f"{session_id}/classification"
    )

    params = {
        "seasonYear": season,
        "test": "false"
    }

    data = get_json(
        url,
        params
    )

    if not data:
        return []

    if isinstance(data, list):

        return data

    if isinstance(data, dict):

        return (
            data.get("classification")
            or data.get("classifications")
            or data.get("content")
            or []
        )

    return []


# ============================================================
# GET GRID
# ============================================================

def get_grid(
    event_uuid,
    category_uuid
):

    url = (
        f"{BASE_URL}/results/event/"
        f"{event_uuid}/category/"
        f"{category_uuid}/grid"
    )

    data = get_json(url)

    if not data:
        return []

    if isinstance(data, list):

        return data

    if isinstance(data, dict):

        return (
            data.get("grid")
            or data.get("content")
            or data.get("results")
            or []
        )

    return []


# ============================================================
# GET RIDER ID
# ============================================================

def get_rider_id(rider):

    if not rider:
        return None

    return (
        rider.get("id")
        or rider.get("uuid")
        or rider.get("legacy_id")
    )


# ============================================================
# DRIVER NAME
# ============================================================

def get_driver_name(rider):

    if not rider:
        return ""

    full_name = (
        rider.get("full_name")
        or rider.get("name")
        or ""
    )

    return str(
        full_name
    ).strip()


# ============================================================
# LAST NAME
# ============================================================

def get_last_name(full_name):

    if not full_name:
        return ""

    parts = full_name.split()

    return parts[-1]


# ============================================================
# DRIVER INITIAL
# ============================================================

def get_driver_initial(full_name):

    if not full_name:
        return ""

    parts = full_name.split()

    if len(parts) == 1:

        return parts[0][0].upper()

    return (
        parts[0][0]
        + parts[-1][0]
    ).upper()


# ============================================================
# GET GRID POSITION
# ============================================================

def get_grid_position(item):

    possible_fields = [

        "grid_position",
        "gridPosition",
        "position",
        "qualifying_position",
        "qualifyingPosition"

    ]

    for field in possible_fields:

        value = item.get(field)

        if value is not None:

            try:

                return int(value)

            except (
                ValueError,
                TypeError
            ):

                pass

    return None


# ============================================================
# CREATE GRID DICTIONARY
# ============================================================

def create_grid_dictionary(grid_data):

    grid_dict = {}

    for item in grid_data:

        rider = (
            item.get("rider")
            or {}
        )

        rider_id = get_rider_id(
            rider
        )

        if rider_id is None:
            continue

        grid_position = (
            get_grid_position(item)
        )

        if grid_position is not None:

            grid_dict[
                str(rider_id)
            ] = grid_position

    return grid_dict


# ============================================================
# GET CLASSIFICATION POSITION
# ============================================================

def get_classification_position(
    result
):

    """
    Mengambil posisi klasifikasi resmi.

    Prioritas:
    1. position
    2. classification_position
    3. classificationPosition
    4. rank
    5. position_number

    DNF / NC / DSQ tetap menggunakan
    posisi klasifikasi apabila API menyediakannya.
    """

    possible_fields = [

        "position",

        "classification_position",

        "classificationPosition",

        "rank",

        "position_number",

        "positionNumber"

    ]

    for field in possible_fields:

        value = result.get(field)

        if value is None:
            continue

        try:

            position = int(value)

            if position > 0:
                return position

        except (
            ValueError,
            TypeError
        ):

            continue

    return None


# ============================================================
# ASSIGN FINISH POSITION
# ============================================================

def assign_finish_positions(
    classification
):

    """
    Mempertahankan posisi resmi dari API.

    Jika posisi kosong, posisi akan diisi
    berdasarkan urutan classification sebagai
    fallback.

    Ini membuat DNF/NC tetap memiliki
    Finish Position numerik apabila API
    tidak memberikan position secara langsung.
    """

    # --------------------------------------------------------
    # Ambil posisi resmi yang tersedia
    # --------------------------------------------------------

    used_positions = set()

    for result in classification:

        position = (
            get_classification_position(
                result
            )
        )

        if position is not None:

            result[
                "_official_finish_position"
            ] = position

            used_positions.add(
                position
            )

        else:

            result[
                "_official_finish_position"
            ] = None

    # --------------------------------------------------------
    # Fallback untuk position kosong
    # --------------------------------------------------------

    next_position = 1

    for result in classification:

        current_position = result.get(
            "_official_finish_position"
        )

        if current_position is not None:
            continue

        while (
            next_position
            in used_positions
        ):

            next_position += 1

        result[
            "_official_finish_position"
        ] = next_position

        used_positions.add(
            next_position
        )

        next_position += 1

    return classification


# ============================================================
# PARSE LAP TIME
# ============================================================

def parse_lap_time(
    lap_time
):

    if lap_time is None:
        return None

    lap_time = str(
        lap_time
    ).strip()

    if not lap_time:
        return None

    try:

        parts = lap_time.split(":")

        # MM:SS.mmm
        if len(parts) == 2:

            minutes = float(
                parts[0]
            )

            seconds = float(
                parts[1]
            )

            return (
                minutes * 60000
                + seconds * 1000
            )

        # HH:MM:SS.mmm
        if len(parts) == 3:

            hours = float(
                parts[0]
            )

            minutes = float(
                parts[1]
            )

            seconds = float(
                parts[2]
            )

            return (
                hours * 3600000
                + minutes * 60000
                + seconds * 1000
            )

    except (
        ValueError,
        TypeError
    ):

        return None

    return None


# ============================================================
# GET BEST LAP
# ============================================================

def get_best_lap_time(result):

    best_lap = (
        result.get("best_lap")
        or result.get("bestLap")
        or {}
    )

    if isinstance(
        best_lap,
        dict
    ):

        return (
            best_lap.get("time")
            or best_lap.get("lap_time")
            or best_lap.get("lapTime")
        )

    if isinstance(
        best_lap,
        str
    ):

        return best_lap

    return None


# ============================================================
# FIND FASTEST LAP
# ============================================================

def find_fastest_lap_rider(
    classification
):

    fastest_rider_id = None

    fastest_time_ms = None

    for result in classification:

        rider = (
            result.get("rider")
            or {}
        )

        rider_id = get_rider_id(
            rider
        )

        if rider_id is None:
            continue

        lap_time = (
            get_best_lap_time(
                result
            )
        )

        lap_time_ms = (
            parse_lap_time(
                lap_time
            )
        )

        if lap_time_ms is None:
            continue

        if (
            fastest_time_ms is None
            or lap_time_ms
            < fastest_time_ms
        ):

            fastest_time_ms = (
                lap_time_ms
            )

            fastest_rider_id = (
                rider_id
            )

    return fastest_rider_id


# ============================================================
# GRAND PRIX NAME
# ============================================================

def get_grand_prix_name(event):

    possible_fields = [

        "name",

        "sponsored_name",

        "sponsoredName",

        "event_name",

        "eventName"

    ]

    for field in possible_fields:

        value = event.get(field)

        if value:

            return str(
                value
            ).strip()

    return ""


# ============================================================
# CIRCUIT NAME
# ============================================================

def get_circuit_name(event):

    circuit = (
        event.get("circuit")
        or {}
    )

    if isinstance(
        circuit,
        dict
    ):

        return (
            circuit.get("name")
            or circuit.get("circuit_name")
            or circuit.get("circuitName")
            or ""
        )

    return (
        event.get("circuit_name")
        or event.get("circuitName")
        or ""
    )


# ============================================================
# CITY
# ============================================================

def get_city(event):

    circuit = (
        event.get("circuit")
        or {}
    )

    if isinstance(
        circuit,
        dict
    ):

        return (
            circuit.get("location")
            or circuit.get("locality")
            or circuit.get("city")
            or ""
        )

    return (
        event.get("city")
        or event.get("location")
        or ""
    )


# ============================================================
# NATION
# ============================================================

def get_nation(event):

    country = event.get(
        "country"
    )

    if isinstance(
        country,
        dict
    ):

        return (
            country.get("name")
            or country.get("iso")
            or country.get("code")
            or ""
        )

    if country:

        return str(
            country
        )

    return (
        event.get("nation")
        or event.get("country_name")
        or ""
    )


# ============================================================
# GP INITIAL
# ============================================================

def get_gp_initial(event):

    possible_fields = [

        "short_name",

        "shortName",

        "code",

        "event_code",

        "eventCode"

    ]

    for field in possible_fields:

        value = event.get(field)

        if value:

            return str(
                value
            ).upper()

    gp_name = get_grand_prix_name(
        event
    )

    if not gp_name:
        return ""

    # Hapus kata umum
    gp_name_clean = re.sub(
        r"\b(GP|GRAND PRIX)\b",
        "",
        gp_name,
        flags=re.IGNORECASE
    ).strip()

    words = (
        gp_name_clean.split()
    )

    if not words:
        return ""

    # Ambil 3 huruf pertama dari
    # kata utama terakhir
    last_word = re.sub(
        r"[^A-Za-z]",
        "",
        words[-1]
    )

    return (
        last_word[:3]
        .upper()
    )


# ============================================================
# GET RACE POINTS
# ============================================================

def get_race_points(result):

    possible_fields = [

        "points",

        "race_points",

        "racePoints",

        "points_scored",

        "pointsScored"

    ]

    for field in possible_fields:

        value = result.get(field)

        if value is not None:

            try:

                return float(value)

            except (
                ValueError,
                TypeError
            ):

                pass

    return 0


# ============================================================
# MAIN PROGRAM
# ============================================================

all_results = []


print()
print("=" * 75)
print("        MOTOGP HISTORICAL RACE DATA SCRAPER")
print("=" * 75)

print(
    f"Season       : {START_SEASON} - {END_SEASON}"
)

print(
    "Race Type    : Main Race"
)

print(
    f"Output       : {OUTPUT_FILE}"
)

print("=" * 75)


# ============================================================
# LOOP SEASON
# ============================================================

for season in range(
    START_SEASON,
    END_SEASON + 1
):

    print()
    print("-" * 75)
    print(
        f"SEASON {season}"
    )
    print("-" * 75)

    # --------------------------------------------------------
    # SEASON ID
    # --------------------------------------------------------

    season_uuid = (
        get_season_id(
            season
        )
    )

    if not season_uuid:

        print(
            f"[ERROR] Season {season} "
            f"tidak ditemukan."
        )

        continue

    print(
        f"Season UUID : {season_uuid}"
    )

    time.sleep(
        REQUEST_DELAY
    )

    # --------------------------------------------------------
    # MOTOGP CATEGORY
    # --------------------------------------------------------

    category_uuid = (
        get_motogp_category_id(
            season_uuid
        )
    )

    if not category_uuid:

        print(
            "[ERROR] Category MotoGP "
            "tidak ditemukan."
        )

        continue

    print(
        f"Category UUID : {category_uuid}"
    )

    time.sleep(
        REQUEST_DELAY
    )

    # --------------------------------------------------------
    # EVENTS
    # --------------------------------------------------------

    events = (
        get_events(
            season_uuid
        )
    )

    if not events:

        print(
            "[ERROR] Event tidak ditemukan."
        )

        continue

    print(
        f"Total Event : {len(events)}"
    )

    # --------------------------------------------------------
    # SORT EVENT
    # --------------------------------------------------------

    events = sorted(
        events,
        key=lambda x: (
            x.get("date")
            or x.get("event_date")
            or x.get("eventDate")
            or ""
        )
    )

    # ========================================================
    # LOOP EVENT
    # ========================================================

    for round_number, event in enumerate(
        events,
        start=1
    ):

        event_uuid = (
            event.get("id")
            or event.get("uuid")
        )

        if not event_uuid:

            continue

        gp_name = (
            get_grand_prix_name(
                event
            )
        )

        print()
        print(
            f"Round {round_number:02d} "
            f"| {gp_name}"
        )

        # ----------------------------------------------------
        # SESSIONS
        # ----------------------------------------------------

        sessions = get_sessions(
            event_uuid,
            category_uuid
        )

        time.sleep(
            REQUEST_DELAY
        )

        if not sessions:

            print(
                "  [SKIP] Session kosong."
            )

            continue

        # ----------------------------------------------------
        # FIND MAIN RACE
        # ----------------------------------------------------

        race_session = (
            find_race_session(
                sessions
            )
        )

        if not race_session:

            print(
                "  [SKIP] Main Race "
                "tidak ditemukan."
            )

            continue

        session_id = (
            race_session.get("id")
            or race_session.get("uuid")
        )

        if not session_id:

            print(
                "  [SKIP] Session ID kosong."
            )

            continue

        print(
            f"  Race Session : {session_id}"
        )

        # ----------------------------------------------------
        # CLASSIFICATION
        # ----------------------------------------------------

        classification = (
            get_classification(
                session_id,
                season
            )
        )

        time.sleep(
            REQUEST_DELAY
        )

        if not classification:

            print(
                "  [SKIP] Classification kosong."
            )

            continue

        # ----------------------------------------------------
        # ASSIGN FINISH POSITION
        # ----------------------------------------------------

        classification = (
            assign_finish_positions(
                classification
            )
        )

        # ----------------------------------------------------
        # GRID
        # ----------------------------------------------------

        grid_data = get_grid(
            event_uuid,
            category_uuid
        )

        time.sleep(
            REQUEST_DELAY
        )

        grid_dict = (
            create_grid_dictionary(
                grid_data
            )
        )

        # ----------------------------------------------------
        # FASTEST LAP
        # ----------------------------------------------------

        fastest_lap_rider_id = (
            find_fastest_lap_rider(
                classification
            )
        )

        # ----------------------------------------------------
        # EVENT INFORMATION
        # ----------------------------------------------------

        circuit_name = (
            get_circuit_name(
                event
            )
        )

        city = (
            get_city(
                event
            )
        )

        nation = (
            get_nation(
                event
            )
        )

        gp_initial = (
            get_gp_initial(
                event
            )
        )

        # ====================================================
        # LOOP DRIVER
        # ====================================================

        for result in classification:

            rider = (
                result.get("rider")
                or {}
            )

            rider_id = get_rider_id(
                rider
            )

            # ------------------------------------------------
            # DRIVER
            # ------------------------------------------------

            driver_name = (
                get_driver_name(
                    rider
                )
            )

            last_name = (
                get_last_name(
                    driver_name
                )
            )

            driver_initial = (
                get_driver_initial(
                    driver_name
                )
            )

            # ------------------------------------------------
            # GRID
            # ------------------------------------------------

            grid_position = None

            if rider_id is not None:

                grid_position = (
                    grid_dict.get(
                        str(rider_id)
                    )
                )

            # ------------------------------------------------
            # FINISH POSITION
            # ------------------------------------------------

            finish_position = (
                result.get(
                    "_official_finish_position"
                )
            )

            # ------------------------------------------------
            # RACE POINTS
            # ------------------------------------------------

            race_points = (
                get_race_points(
                    result
                )
            )

            # ------------------------------------------------
            # FASTEST LAP
            # ------------------------------------------------

            if (
                rider_id is not None
                and fastest_lap_rider_id is not None
                and str(rider_id)
                == str(
                    fastest_lap_rider_id
                )
            ):

                fastest_lap = 1

            else:

                fastest_lap = 0

            # ------------------------------------------------
            # CREATE ROW
            # ------------------------------------------------

            row = {

                "Round":
                    round_number,

                "Season":
                    season,

                "IsLatestSeason":
                    (
                        1
                        if season == END_SEASON
                        else 0
                    ),

                "Grand Prix":
                    gp_name,

                "Circuit Name":
                    circuit_name,

                "City":
                    city,

                "Nation":
                    nation,

                "GP Initial":
                    gp_initial,

                "Race Type":
                    "Main Race",

                "Driver Name":
                    driver_name,

                "Last Name":
                    last_name,

                "Driver Initial":
                    driver_initial,

                "Grid Position":
                    grid_position,

                "Finish Position":
                    finish_position,

                "Race Points":
                    race_points,

                "Fastest Lap":
                    fastest_lap
            }

            all_results.append(
                row
            )

        print(
            f"  Drivers       : "
            f"{len(classification)}"
        )

        print(
            f"  Fastest Lap   : "
            f"{fastest_lap_rider_id}"
        )


# ============================================================
# CREATE DATAFRAME
# ============================================================

print()
print("=" * 75)
print("MEMBUAT DATAFRAME")
print("=" * 75)


df = pd.DataFrame(
    all_results
)


# ============================================================
# COLUMN ORDER
# ============================================================

columns = [

    "Round",
    "Season",
    "IsLatestSeason",
    "Grand Prix",
    "Circuit Name",
    "City",
    "Nation",
    "GP Initial",
    "Race Type",
    "Driver Name",
    "Last Name",
    "Driver Initial",
    "Grid Position",
    "Finish Position",
    "Race Points",
    "Fastest Lap"

]

if not df.empty:

    df = df[columns]


# ============================================================
# DATA CLEANING
# ============================================================

if not df.empty:

    # Round
    df["Round"] = pd.to_numeric(
        df["Round"],
        errors="coerce"
    ).astype("Int64")

    # Season
    df["Season"] = pd.to_numeric(
        df["Season"],
        errors="coerce"
    ).astype("Int64")

    # IsLatestSeason
    df["IsLatestSeason"] = (
        pd.to_numeric(
            df["IsLatestSeason"],
            errors="coerce"
        )
        .fillna(0)
        .astype(int)
    )

    # Grid
    df["Grid Position"] = (
        pd.to_numeric(
            df["Grid Position"],
            errors="coerce"
        )
        .astype("Int64")
    )

    # Finish
    df["Finish Position"] = (
        pd.to_numeric(
            df["Finish Position"],
            errors="coerce"
        )
        .astype("Int64")
    )

    # Points
    df["Race Points"] = (
        pd.to_numeric(
            df["Race Points"],
            errors="coerce"
        )
        .fillna(0)
    )

    # Fastest Lap
    df["Fastest Lap"] = (
        pd.to_numeric(
            df["Fastest Lap"],
            errors="coerce"
        )
        .fillna(0)
        .astype(int)
    )

    # --------------------------------------------------------
    # SORT
    # --------------------------------------------------------

    df = df.sort_values(
        by=[
            "Season",
            "Round",
            "Finish Position"
        ],
        ascending=[
            True,
            True,
            True
        ],
        na_position="last"
    )

    df = df.reset_index(
        drop=True
    )


# ============================================================
# SAVE EXCEL
# ============================================================

print()
print("=" * 75)
print("MENYIMPAN EXCEL")
print("=" * 75)


with pd.ExcelWriter(
    OUTPUT_FILE,
    engine="openpyxl"
) as writer:

    df.to_excel(
        writer,
        sheet_name="Driver",
        index=False
    )

    worksheet = writer.sheets[
        "Driver"
    ]

    # --------------------------------------------------------
    # FREEZE HEADER
    # --------------------------------------------------------

    worksheet.freeze_panes = "A2"

    # --------------------------------------------------------
    # AUTO FILTER
    # --------------------------------------------------------

    if worksheet.max_row > 1:

        worksheet.auto_filter.ref = (
            worksheet.dimensions
        )

    # --------------------------------------------------------
    # AUTO WIDTH
    # --------------------------------------------------------

    for column_cells in worksheet.columns:

        max_length = 0

        column_index = (
            column_cells[0].column
        )

        column_letter = (
            get_column_letter(
                column_index
            )
        )

        for cell in column_cells:

            try:

                value = str(
                    cell.value
                )

                if len(value) > max_length:

                    max_length = len(
                        value
                    )

            except Exception:

                pass

        worksheet.column_dimensions[
            column_letter
        ].width = min(
            max_length + 2,
            35
        )


# ============================================================
# SUMMARY
# ============================================================

print()
print("=" * 75)
print("SELESAI")
print("=" * 75)

print(
    f"Total Row        : {len(df):,}"
)

if not df.empty:

    total_races = (
        df[
            [
                "Season",
                "Round"
            ]
        ]
        .drop_duplicates()
        .shape[0]
    )

    total_drivers = (
        df["Driver Name"]
        .nunique()
    )

    print(
        f"Total Race       : {total_races:,}"
    )

    print(
        f"Total Driver     : {total_drivers:,}"
    )

print(
    f"Output File      : {OUTPUT_FILE}"
)

print(
    "Sheet            : Driver"
)

print("=" * 75)