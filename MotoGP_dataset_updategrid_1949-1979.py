# ============================================================
# MOTOGP / 500CC HISTORICAL DATA SCRAPER
# ============================================================
#
# Data:
# - Premier Class
# - 500cc    : <= 2001
# - MotoGP   : >= 2002
#
# Range default:
# - 1990 sampai 1999
#
# Output:
# - Excel
# - Sheet Driver
# - Sheet Unavailable_Season
#
# Grid Position Fallback:
# 1. Official Grid API
# 2. Race Classification
# 3. Qualifying Classification
# 4. None
#
# ============================================================


import requests
import pandas as pd
import time
import re

from openpyxl.utils import get_column_letter


# ============================================================
# CONFIGURATION
# ============================================================

START_SEASON = 1949
END_SEASON = 1979


OUTPUT_FILE = (
    f"MotoGP_Historical_"
    f"{START_SEASON}_{END_SEASON}.xlsx"
)


BASE_URL = (
    "https://api.motogp.pulselive.com/motogp/v1"
)


REQUEST_DELAY = 0.3
TIMEOUT = 30
MAX_RETRIES = 3


HEADERS = {

    "User-Agent": (
        "Mozilla/5.0 "
        "(Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 "
        "(KHTML, like Gecko) "
        "Chrome/120.0.0.0 "
        "Safari/537.36"
    ),

    "Accept": "application/json",

    "Origin": "https://www.motogp.com",

    "Referer": (
        "https://www.motogp.com/"
    )

}


# ============================================================
# GET JSON
# ============================================================

def get_json(
    url,
    params=None,
    retries=MAX_RETRIES
):

    for attempt in range(
        1,
        retries + 1
    ):

        try:

            response = requests.get(

                url,

                params=params,

                headers=HEADERS,

                timeout=TIMEOUT

            )


            response.raise_for_status()


            return response.json()


        except requests.exceptions.RequestException as error:

            print(

                f"Request gagal "
                f"({attempt}/{retries}): "
                f"{error}"

            )


            if attempt < retries:

                time.sleep(2)


    return None


# ============================================================
# NORMALIZE API LIST
# ============================================================

def extract_list(
    data,
    possible_keys
):

    if data is None:

        return []


    if isinstance(
        data,
        list
    ):

        return data


    if isinstance(
        data,
        dict
    ):

        for key in possible_keys:

            value = data.get(key)


            if isinstance(
                value,
                list
            ):

                return value


    return []


# ============================================================
# GET ALL SEASONS
# ============================================================

def get_all_seasons():

    url = (
        f"{BASE_URL}/results/seasons"
    )


    data = get_json(
        url
    )


    seasons = extract_list(

        data,

        [

            "seasons",

            "content",

            "results"

        ]

    )


    return seasons


# ============================================================
# CREATE SEASON DICTIONARY
# ============================================================

def create_season_dictionary(
    seasons
):

    season_dict = {}


    for item in seasons:


        year = item.get(
            "year"
        )


        season_id = (

            item.get("id")

            or item.get("uuid")

        )


        if (

            year is None

            or season_id is None

        ):

            continue


        try:

            year = int(
                year
            )


            season_dict[
                year
            ] = season_id


        except (

            ValueError,

            TypeError

        ):

            continue


    return season_dict


# ============================================================
# GET CATEGORIES
# ============================================================

def get_categories(
    season_uuid
):

    url = (
        f"{BASE_URL}/results/categories"
    )


    params = {

        "seasonUuid":
            season_uuid

    }


    data = get_json(

        url,

        params

    )


    categories = extract_list(

        data,

        [

            "categories",

            "content",

            "results"

        ]

    )


    return categories


# ============================================================
# FIND PREMIER CATEGORY
# ============================================================

def find_premier_category(
    categories,
    season
):

    if season <= 2001:


        keywords = [

            "500cc",

            "500 cc",

            "500"

        ]


    else:


        keywords = [

            "motogp"

        ]


    for category in categories:


        category_name = str(

            category.get(

                "name",

                ""

            )

        ).lower().strip()


        for keyword in keywords:


            if keyword in category_name:


                category_id = (

                    category.get("id")

                    or category.get("uuid")

                )


                if category_id:

                    return category_id


    return None


# ============================================================
# GET EVENTS
# ============================================================

def get_events(
    season_uuid
):

    url = (
        f"{BASE_URL}/results/events"
    )


    params = {

        "seasonUuid":
            season_uuid,

        "isFinished":
            "true"

    }


    data = get_json(

        url,

        params

    )


    events = extract_list(

        data,

        [

            "events",

            "content",

            "results"

        ]

    )


    return events


# ============================================================
# SORT EVENTS
# ============================================================

def sort_events(
    events
):

    def get_event_sort_value(
        event
    ):

        return (

            event.get("sequence")

            or event.get("round")

            or event.get("roundNumber")

            or event.get("date")

            or event.get("eventDate")

            or ""

        )


    return sorted(

        events,

        key=get_event_sort_value

    )


# ============================================================
# GET EVENT ID
# ============================================================

def get_event_id(
    event
):

    return (

        event.get("id")

        or event.get("uuid")

    )


# ============================================================
# GET SESSIONS
# ============================================================

def get_sessions(

    event_uuid,

    category_uuid

):

    url = (
        f"{BASE_URL}/results/sessions"
    )


    params = {

        "eventUuid":
            event_uuid,

        "categoryUuid":
            category_uuid

    }


    data = get_json(

        url,

        params

    )


    sessions = extract_list(

        data,

        [

            "sessions",

            "content",

            "results"

        ]

    )


    return sessions


# ============================================================
# FIND MAIN RACE
# ============================================================

def find_main_race(
    sessions
):

    for session in sessions:


        session_type = str(

            session.get(

                "type",

                ""

            )

        ).upper().strip()


        if session_type == "RAC":

            return session


    for session in sessions:


        session_name = str(

            session.get(

                "name",

                ""

            )

        ).upper().strip()


        if (

            session_name == "RACE"

            or session_name == "MAIN RACE"

            or "RACE" == session_name

        ):

            return session


    return None


# ============================================================
# FIND QUALIFYING SESSION
# ============================================================

def find_qualifying_session(
    sessions
):

    qualifying_types = [

        "Q1",

        "Q2",

        "QP",

        "QUA",

        "QUALIFYING"

    ]


    # --------------------------------------------------------
    # PRIORITAS SESSION TYPE
    # --------------------------------------------------------

    for session in sessions:


        session_type = str(

            session.get(

                "type",

                ""

            )

        ).upper().strip()


        if session_type in qualifying_types:

            return session


    # --------------------------------------------------------
    # PRIORITAS SESSION NAME
    # --------------------------------------------------------

    for session in sessions:


        session_name = str(

            session.get(

                "name",

                ""

            )

        ).upper().strip()


        if (

            "QUALIFY"

            in session_name

        ):

            return session


    return None


# ============================================================
# GET SESSION ID
# ============================================================

def get_session_id(
    session
):

    return (

        session.get("id")

        or session.get("uuid")

    )


# ============================================================
# GET CLASSIFICATION
# ============================================================

def get_classification(
    session_id
):

    url = (
        f"{BASE_URL}/results/session/"
        f"{session_id}/classification"
    )


    params = {

        "test":
            "false"

    }


    data = get_json(

        url,

        params

    )


    classification = extract_list(

        data,

        [

            "classification",

            "classifications",

            "content",

            "results"

        ]

    )


    return classification


# ============================================================
# GET GRID DATA
# ============================================================

def get_grid(

    event_uuid,

    category_uuid

):

    url = (
        f"{BASE_URL}/results/event/"
        f"{event_uuid}/category/"
        f"{category_uuid}/grid"
    )


    data = get_json(
        url
    )


    grid = extract_list(

        data,

        [

            "grid",

            "content",

            "results"

        ]

    )


    return grid


# ============================================================
# GET RIDER ID
# ============================================================

def get_rider_id(
    rider
):

    if not rider:

        return None


    return (

        rider.get("id")

        or rider.get("uuid")

        or rider.get("legacy_id")

    )


# ============================================================
# GET DRIVER NAME
# ============================================================

def get_driver_name(
    rider
):

    if not rider:

        return ""


    possible_fields = [

        "full_name",

        "fullName",

        "name"

    ]


    for field in possible_fields:


        value = rider.get(
            field
        )


        if value:

            return str(
                value
            ).strip()


    first_name = (

        rider.get("first_name")

        or rider.get("firstName")

        or ""

    )


    last_name = (

        rider.get("last_name")

        or rider.get("lastName")

        or ""

    )


    return (

        f"{first_name} "
        f"{last_name}"

    ).strip()


# ============================================================
# GET LAST NAME
# ============================================================

def get_last_name(

    rider,

    driver_name

):

    if rider:


        last_name = (

            rider.get("last_name")

            or rider.get("lastName")

        )


        if last_name:

            return str(
                last_name
            ).strip()


    if not driver_name:

        return ""


    return (

        driver_name
        .split()[-1]

    )


# ============================================================
# GET DRIVER INITIAL
# ============================================================

def get_driver_initial(

    rider,

    driver_name

):

    if rider:


        short_name = (

            rider.get("short_name")

            or rider.get("shortName")

            or rider.get("abbreviation")

        )


        if short_name:

            return str(
                short_name
            ).upper()


    if not driver_name:

        return ""


    parts = (
        driver_name
        .split()
    )


    if len(parts) == 1:

        return (
            parts[0][0]
            .upper()
        )


    return (

        parts[0][0]

        + parts[-1][0]

    ).upper()


# ============================================================
# GET NUMERIC POSITION
# ============================================================

def get_numeric_position(
    item
):

    possible_fields = [

        "position",

        "classification_position",

        "classificationPosition",

        "rank",

        "position_number",

        "positionNumber"

    ]


    for field in possible_fields:


        value = item.get(
            field
        )


        if value is None:

            continue


        try:


            position = int(
                float(value)
            )


            if position > 0:

                return position


        except (

            ValueError,

            TypeError

        ):

            pass


    return None


# ============================================================
# GET GRID POSITION FROM RESULT
# ============================================================

def get_grid_from_result(
    result
):

    possible_fields = [

        "grid_position",

        "gridPosition",

        "start_position",

        "startPosition",

        "starting_position",

        "startingPosition",

        "grid",

        "start"

    ]


    for field in possible_fields:


        value = result.get(
            field
        )


        if value is None:

            continue


        try:


            position = int(
                float(value)
            )


            if position > 0:

                return position


        except (

            ValueError,

            TypeError

        ):

            pass


    return None


# ============================================================
# ASSIGN FINISH POSITIONS
# ============================================================

def assign_finish_positions(
    classification
):

    """
    Prioritas:

    1. Posisi resmi dari API

    2. Jika posisi kosong,
       gunakan urutan classification
       dari API.

    Catatan:
    DNF / DNS / DSQ tetap menggunakan
    posisi numerik jika API menyediakannya.
    """


    used_positions = set()


    # --------------------------------------------------------
    # PASS 1
    # --------------------------------------------------------

    for result in classification:


        position = (

            get_numeric_position(
                result
            )

        )


        result[
            "_finish_position"
        ] = position


        if position is not None:

            used_positions.add(
                position
            )


    # --------------------------------------------------------
    # PASS 2
    # --------------------------------------------------------

    next_position = 1


    for result in classification:


        if (

            result.get(
                "_finish_position"
            )

            is not None

        ):

            continue


        while (

            next_position

            in used_positions

        ):

            next_position += 1


        result[
            "_finish_position"
        ] = next_position


        used_positions.add(
            next_position
        )


        next_position += 1


    return classification


# ============================================================
# CREATE GRID DICTIONARY
# ============================================================

def create_grid_dictionary(
    grid_data
):

    grid_dict = {}


    for item in grid_data:


        rider = (

            item.get("rider")

            or {}

        )


        rider_id = (
            get_rider_id(
                rider
            )
        )


        if rider_id is None:

            continue


        possible_fields = [

            "grid_position",

            "gridPosition",

            "start_position",

            "startPosition",

            "starting_position",

            "startingPosition",

            "position",

            "qualifying_position",

            "qualifyingPosition"

        ]


        grid_position = None


        for field in possible_fields:


            value = item.get(
                field
            )


            if value is None:

                continue


            try:


                position = int(
                    float(value)
                )


                if position > 0:


                    grid_position = (
                        position
                    )


                    break


            except (

                ValueError,

                TypeError

            ):

                pass


        if grid_position is not None:


            grid_dict[
                str(
                    rider_id
                )
            ] = grid_position


    return grid_dict


# ============================================================
# CREATE QUALIFYING GRID DICTIONARY
# ============================================================

def create_qualifying_grid_dictionary(
    qualifying_classification
):

    qualifying_dict = {}


    for result in qualifying_classification:


        rider = (

            result.get("rider")

            or {}

        )


        rider_id = (
            get_rider_id(
                rider
            )
        )


        if rider_id is None:

            continue


        qualifying_position = (

            get_numeric_position(
                result
            )

        )


        if qualifying_position is not None:


            qualifying_dict[
                str(
                    rider_id
                )
            ] = qualifying_position


    return qualifying_dict


# ============================================================
# GET RACE POINTS
# ============================================================

def get_race_points(
    result
):

    possible_fields = [

        "points",

        "race_points",

        "racePoints",

        "points_scored",

        "pointsScored"

    ]


    for field in possible_fields:


        value = result.get(
            field
        )


        if value is None:

            continue


        try:

            return float(
                value
            )


        except (

            ValueError,

            TypeError

        ):

            pass


    return 0


# ============================================================
# PARSE LAP TIME
# ============================================================

def parse_lap_time(
    value
):

    if value is None:

        return None


    value = str(
        value
    ).strip()


    if not value:

        return None


    try:


        parts = value.split(
            ":"
        )


        # MM:SS.xxx

        if len(parts) == 2:


            minutes = float(
                parts[0]
            )


            seconds = float(
                parts[1]
            )


            return (

                minutes * 60000

                + seconds * 1000

            )


        # HH:MM:SS.xxx

        elif len(parts) == 3:


            hours = float(
                parts[0]
            )


            minutes = float(
                parts[1]
            )


            seconds = float(
                parts[2]
            )


            return (

                hours * 3600000

                + minutes * 60000

                + seconds * 1000

            )


    except (

        ValueError,

        TypeError

    ):

        return None


    return None


# ============================================================
# GET BEST LAP TIME
# ============================================================

def get_best_lap_time(
    result
):

    best_lap = (

        result.get("best_lap")

        or result.get("bestLap")

        or {}

    )


    if isinstance(

        best_lap,

        dict

    ):

        return (

            best_lap.get("time")

            or best_lap.get("lap_time")

            or best_lap.get("lapTime")

        )


    if isinstance(

        best_lap,

        str

    ):

        return best_lap


    return None


# ============================================================
# FIND FASTEST LAP RIDER
# ============================================================

def find_fastest_lap_rider(
    classification
):

    fastest_rider_id = None

    fastest_time = None


    for result in classification:


        rider = (

            result.get("rider")

            or {}

        )


        rider_id = (
            get_rider_id(
                rider
            )
        )


        if rider_id is None:

            continue


        lap_time = (

            get_best_lap_time(
                result
            )

        )


        lap_time_ms = (

            parse_lap_time(
                lap_time
            )

        )


        if lap_time_ms is None:

            continue


        if (

            fastest_time is None

            or lap_time_ms < fastest_time

        ):


            fastest_time = (
                lap_time_ms
            )


            fastest_rider_id = (
                rider_id
            )


    return fastest_rider_id


# ============================================================
# GET GRAND PRIX NAME
# ============================================================

def get_grand_prix_name(
    event
):

    possible_fields = [

        "name",

        "sponsored_name",

        "sponsoredName",

        "event_name",

        "eventName"

    ]


    for field in possible_fields:


        value = event.get(
            field
        )


        if value:

            return str(
                value
            ).strip()


    return ""


# ============================================================
# GET CIRCUIT NAME
# ============================================================

def get_circuit_name(
    event
):

    circuit = (

        event.get("circuit")

        or {}

    )


    if isinstance(

        circuit,

        dict

    ):


        possible_fields = [

            "name",

            "circuit_name",

            "circuitName"

        ]


        for field in possible_fields:


            value = circuit.get(
                field
            )


            if value:

                return str(
                    value
                ).strip()


    return (

        event.get("circuit_name")

        or event.get("circuitName")

        or ""

    )


# ============================================================
# GET CITY
# ============================================================

def get_city(
    event
):

    circuit = (

        event.get("circuit")

        or {}

    )


    if isinstance(

        circuit,

        dict

    ):


        possible_fields = [

            "location",

            "locality",

            "city",

            "town"

        ]


        for field in possible_fields:


            value = circuit.get(
                field
            )


            if value:

                return str(
                    value
                ).strip()


    return (

        event.get("city")

        or event.get("location")

        or ""

    )


# ============================================================
# GET NATION
# ============================================================

def get_nation(
    event
):

    country = event.get(
        "country"
    )


    if isinstance(

        country,

        dict

    ):

        return (

            country.get("name")

            or country.get("country_name")

            or country.get("iso")

            or ""

        )


    if country:

        return str(
            country
        )


    return (

        event.get("nation")

        or event.get("country_name")

        or ""

    )


# ============================================================
# GET GP INITIAL
# ============================================================

def get_gp_initial(
    event
):

    possible_fields = [

        "short_name",

        "shortName",

        "code",

        "event_code",

        "eventCode"

    ]


    for field in possible_fields:


        value = event.get(
            field
        )


        if value:

            return str(
                value
            ).upper()


    # --------------------------------------------------------
    # FALLBACK
    # --------------------------------------------------------

    gp_name = (

        get_grand_prix_name(
            event
        )

    )


    if not gp_name:

        return ""


    clean_name = re.sub(

        r"\b(GP|GRAND PRIX)\b",

        "",

        gp_name,

        flags=re.IGNORECASE

    ).strip()


    words = (
        clean_name
        .split()
    )


    if not words:

        return ""


    word = re.sub(

        r"[^A-Za-z]",

        "",

        words[-1]

    )


    return (

        word[:3]

        .upper()

    )


# ============================================================
# GET PREMIER CLASS NAME
# ============================================================

def get_premier_class_name(
    season
):

    if season <= 2001:

        return "500cc"


    return "MotoGP"


# ============================================================
# MAIN PROGRAM
# ============================================================

def main():


    all_results = []

    unavailable_seasons = []


    print()

    print(
        "=" * 80
    )

    print(
        "MOTOGP / 500CC "
        "HISTORICAL SCRAPER"
    )

    print(
        "=" * 80
    )


    print(

        f"Range Season : "
        f"{START_SEASON} - {END_SEASON}"

    )


    print(

        f"Output       : "
        f"{OUTPUT_FILE}"

    )


    print(
        "=" * 80
    )


    # ========================================================
    # GET SEASONS
    # ========================================================

    print()

    print(
        "Mengambil daftar season..."
    )


    seasons = (
        get_all_seasons()
    )


    if not seasons:


        print()

        print(
            "ERROR:"
        )


        print(

            "Tidak dapat mengambil "
            "daftar season dari API."

        )


        return


    season_dict = (

        create_season_dictionary(
            seasons
        )

    )


    print(

        f"Total season tersedia: "
        f"{len(season_dict)}"

    )


    # ========================================================
    # LOOP SEASON
    # ========================================================

    for season in range(

        START_SEASON,

        END_SEASON + 1

    ):


        print()

        print(
            "-" * 80
        )


        print(

            f"SEASON {season} "

            f"| "

            f"{get_premier_class_name(season)}"

        )


        print(
            "-" * 80
        )


        # ====================================================
        # GET SEASON UUID
        # ====================================================

        season_uuid = (

            season_dict.get(
                season
            )

        )


        if season_uuid is None:


            print(

                "[UNAVAILABLE] "
                "Season tidak tersedia."

            )


            unavailable_seasons.append({

                "Season":
                    season,

                "Class":
                    get_premier_class_name(
                        season
                    ),

                "Reason":
                    (
                        "Season tidak tersedia "
                        "di API"
                    )

            })


            continue


        # ====================================================
        # GET CATEGORIES
        # ====================================================

        categories = (

            get_categories(
                season_uuid
            )

        )


        time.sleep(
            REQUEST_DELAY
        )


        if not categories:


            print(

                "[UNAVAILABLE] "
                "Category kosong."

            )


            unavailable_seasons.append({

                "Season":
                    season,

                "Class":
                    get_premier_class_name(
                        season
                    ),

                "Reason":
                    (
                        "Category tidak ditemukan"
                    )

            })


            continue


        # ====================================================
        # FIND PREMIER CATEGORY
        # ====================================================

        category_uuid = (

            find_premier_category(

                categories,

                season

            )

        )


        if category_uuid is None:


            available_categories = (

                ", ".join(

                    str(

                        item.get(
                            "name",
                            ""
                        )

                    )

                    for item in categories

                )

            )


            print(

                "[UNAVAILABLE] "
                "Premier category "
                "tidak ditemukan."

            )


            print(

                f"Category tersedia: "
                f"{available_categories}"

            )


            unavailable_seasons.append({

                "Season":
                    season,

                "Class":
                    get_premier_class_name(
                        season
                    ),

                "Reason":
                    (
                        "Premier category "
                        "tidak ditemukan"
                    )

            })


            continue


        print(

            f"Category UUID: "
            f"{category_uuid}"

        )


        # ====================================================
        # GET EVENTS
        # ====================================================

        events = (

            get_events(
                season_uuid
            )

        )


        time.sleep(
            REQUEST_DELAY
        )


        if not events:


            print(

                "[UNAVAILABLE] "
                "Event tidak ditemukan."

            )


            unavailable_seasons.append({

                "Season":
                    season,

                "Class":
                    get_premier_class_name(
                        season
                    ),

                "Reason":
                    (
                        "Event tidak ditemukan"
                    )

            })


            continue


        events = (
            sort_events(
                events
            )
        )


        print(

            f"Total Event: "
            f"{len(events)}"

        )


        season_rows_before = (
            len(all_results)
        )


        # ====================================================
        # LOOP EVENT
        # ====================================================

        for round_number, event in enumerate(

            events,

            start=1

        ):


            event_uuid = (

                get_event_id(
                    event
                )

            )


            if event_uuid is None:


                print(

                    f"Round {round_number}: "
                    "Event UUID kosong."

                )


                continue


            grand_prix = (

                get_grand_prix_name(
                    event
                )

            )


            print()


            print(

                f"Round "
                f"{round_number:02d} "

                f"| "

                f"{grand_prix}"

            )


            # =================================================
            # GET SESSIONS
            # =================================================

            sessions = (

                get_sessions(

                    event_uuid,

                    category_uuid

                )

            )


            time.sleep(
                REQUEST_DELAY
            )


            if not sessions:


                print(

                    "  SKIP: "
                    "Session tidak ditemukan."

                )


                continue


            # =================================================
            # FIND MAIN RACE
            # =================================================

            race_session = (

                find_main_race(
                    sessions
                )

            )


            if race_session is None:


                print(

                    "  SKIP: "
                    "Main Race tidak ditemukan."

                )


                continue


            session_id = (

                get_session_id(
                    race_session
                )

            )


            if session_id is None:


                print(

                    "  SKIP: "
                    "Race Session ID kosong."

                )


                continue


            # =================================================
            # FIND QUALIFYING
            # =================================================

            qualifying_session = (

                find_qualifying_session(
                    sessions
                )

            )


            # =================================================
            # GET RACE CLASSIFICATION
            # =================================================

            classification = (

                get_classification(
                    session_id
                )

            )


            time.sleep(
                REQUEST_DELAY
            )


            if not classification:


                print(

                    "  SKIP: "
                    "Race Classification kosong."

                )


                continue


            # =================================================
            # ASSIGN FINISH POSITION
            # =================================================

            classification = (

                assign_finish_positions(
                    classification
                )

            )


            # =================================================
            # GET OFFICIAL GRID
            # =================================================

            grid_data = (

                get_grid(

                    event_uuid,

                    category_uuid

                )

            )


            time.sleep(
                REQUEST_DELAY
            )


            grid_dict = (

                create_grid_dictionary(
                    grid_data
                )

            )


            # =================================================
            # GET QUALIFYING CLASSIFICATION
            # =================================================

            qualifying_classification = []


            if qualifying_session:


                qualifying_session_id = (

                    get_session_id(
                        qualifying_session
                    )

                )


                if qualifying_session_id:


                    qualifying_classification = (

                        get_classification(
                            qualifying_session_id
                        )

                    )


                    time.sleep(
                        REQUEST_DELAY
                    )


            # =================================================
            # CREATE QUALIFYING GRID DICTIONARY
            # =================================================

            qualifying_grid_dict = (

                create_qualifying_grid_dictionary(

                    qualifying_classification

                )

            )


            # =================================================
            # FIND FASTEST LAP
            # =================================================

            fastest_lap_rider_id = (

                find_fastest_lap_rider(
                    classification
                )

            )


            # =================================================
            # EVENT INFORMATION
            # =================================================

            circuit_name = (

                get_circuit_name(
                    event
                )

            )


            city = (

                get_city(
                    event
                )

            )


            nation = (

                get_nation(
                    event
                )

            )


            gp_initial = (

                get_gp_initial(
                    event
                )

            )


            # =================================================
            # LOOP DRIVER
            # =================================================

            for result in classification:


                # ------------------------------------------------
                # RIDER
                # ------------------------------------------------

                rider = (

                    result.get("rider")

                    or {}

                )


                rider_id = (

                    get_rider_id(
                        rider
                    )

                )


                # ------------------------------------------------
                # DRIVER NAME
                # ------------------------------------------------

                driver_name = (

                    get_driver_name(
                        rider
                    )

                )


                # ------------------------------------------------
                # LAST NAME
                # ------------------------------------------------

                last_name = (

                    get_last_name(

                        rider,

                        driver_name

                    )

                )


                # ------------------------------------------------
                # DRIVER INITIAL
                # ------------------------------------------------

                driver_initial = (

                    get_driver_initial(

                        rider,

                        driver_name

                    )

                )


                # =================================================
                # GRID POSITION
                # =================================================

                grid_position = None


                # ------------------------------------------------
                # PRIORITY 1
                # OFFICIAL GRID API
                # ------------------------------------------------

                if rider_id is not None:


                    grid_position = (

                        grid_dict.get(

                            str(
                                rider_id
                            )

                        )

                    )


                # ------------------------------------------------
                # PRIORITY 2
                # RACE CLASSIFICATION
                # ------------------------------------------------

                if grid_position is None:


                    grid_position = (

                        get_grid_from_result(
                            result
                        )

                    )


                # ------------------------------------------------
                # PRIORITY 3
                # QUALIFYING CLASSIFICATION
                # ------------------------------------------------

                if (

                    grid_position is None

                    and rider_id is not None

                ):


                    grid_position = (

                        qualifying_grid_dict.get(

                            str(
                                rider_id
                            )

                        )

                    )


                # =================================================
                # FINISH POSITION
                # =================================================

                finish_position = (

                    result.get(
                        "_finish_position"
                    )

                )


                # =================================================
                # RACE POINTS
                # =================================================

                race_points = (

                    get_race_points(
                        result
                    )

                )


                # =================================================
                # FASTEST LAP
                # =================================================

                fastest_lap = 0


                if (

                    rider_id is not None

                    and fastest_lap_rider_id
                    is not None

                    and str(
                        rider_id
                    )

                    == str(
                        fastest_lap_rider_id
                    )

                ):


                    fastest_lap = 1


                # =================================================
                # SAVE RESULT
                # =================================================

                all_results.append({


                    "Round":
                        round_number,


                    "Season":
                        season,


                    "IsLatestSeason":

                        1

                        if season == END_SEASON

                        else 0,


                    "Grand Prix":
                        grand_prix,


                    "Circuit Name":
                        circuit_name,


                    "City":
                        city,


                    "Nation":
                        nation,


                    "GP Initial":
                        gp_initial,


                    "Race Type":
                        "Main Race",


                    "Driver Name":
                        driver_name,


                    "Last Name":
                        last_name,


                    "Driver Initial":
                        driver_initial,


                    "Grid Position":
                        grid_position,


                    "Finish Position":
                        finish_position,


                    "Race Points":
                        race_points,


                    "Fastest Lap":
                        fastest_lap

                })


            print(

                f"  Driver ditemukan: "
                f"{len(classification)}"

            )


            print(

                f"  Grid Official: "
                f"{len(grid_dict)}"

            )


            print(

                f"  Grid Qualifying: "
                f"{len(qualifying_grid_dict)}"

            )


        # ====================================================
        # SEASON SUMMARY
        # ====================================================

        season_rows_after = (
            len(all_results)
        )


        season_total_rows = (

            season_rows_after

            - season_rows_before

        )


        print()


        print(

            f"SEASON {season} "
            f"SELESAI"

        )


        print(

            f"Total Driver Rows: "
            f"{season_total_rows}"

        )


    # ========================================================
    # CREATE DATAFRAME
    # ========================================================

    print()

    print(
        "=" * 80
    )

    print(
        "MEMBUAT DATAFRAME"
    )

    print(
        "=" * 80
    )


    columns = [


        "Round",


        "Season",


        "IsLatestSeason",


        "Grand Prix",


        "Circuit Name",


        "City",


        "Nation",


        "GP Initial",


        "Race Type",


        "Driver Name",


        "Last Name",


        "Driver Initial",


        "Grid Position",


        "Finish Position",


        "Race Points",


        "Fastest Lap"

    ]


    df = pd.DataFrame(
        all_results
    )


    if not df.empty:


        df = df[
            columns
        ]


        # ====================================================
        # NUMERIC CONVERSION
        # ====================================================

        numeric_columns = [


            "Round",


            "Season",


            "IsLatestSeason",


            "Grid Position",


            "Finish Position",


            "Race Points",


            "Fastest Lap"

        ]


        for column in numeric_columns:


            df[column] = pd.to_numeric(

                df[column],

                errors="coerce"

            )


        # ====================================================
        # INTEGER
        # ====================================================

        integer_columns = [


            "Round",


            "Season",


            "IsLatestSeason",


            "Grid Position",


            "Finish Position",


            "Fastest Lap"

        ]


        for column in integer_columns:


            df[column] = (

                df[column]

                .astype(
                    "Int64"
                )

            )


        # ====================================================
        # SORT DATA
        # ====================================================

        df = df.sort_values(

            by=[

                "Season",

                "Round",

                "Finish Position"

            ],

            ascending=[

                True,

                True,

                True

            ],

            na_position="last"

        )


        df = df.reset_index(
            drop=True
        )


    # ========================================================
    # UNAVAILABLE SEASON DATAFRAME
    # ========================================================

    unavailable_df = pd.DataFrame(
        unavailable_seasons
    )


    if unavailable_df.empty:


        unavailable_df = pd.DataFrame({

            "Season": [],

            "Class": [],

            "Reason": []

        })


    # ========================================================
    # SAVE TO EXCEL
    # ========================================================

    print()

    print(
        "=" * 80
    )

    print(
        "MENYIMPAN FILE EXCEL"
    )

    print(
        "=" * 80
    )


    with pd.ExcelWriter(

        OUTPUT_FILE,

        engine="openpyxl"

    ) as writer:


        # ====================================================
        # DRIVER SHEET
        # ====================================================

        df.to_excel(

            writer,

            sheet_name="Driver",

            index=False

        )


        # ====================================================
        # UNAVAILABLE SEASON SHEET
        # ====================================================

        unavailable_df.to_excel(

            writer,

            sheet_name="Unavailable_Season",

            index=False

        )


        # ====================================================
        # FORMAT WORKSHEET
        # ====================================================

        for sheet_name in [


            "Driver",


            "Unavailable_Season"

        ]:


            worksheet = (

                writer.sheets[
                    sheet_name
                ]

            )


            # Freeze Header

            worksheet.freeze_panes = (
                "A2"
            )


            # Auto Filter

            if worksheet.max_row > 1:


                worksheet.auto_filter.ref = (

                    worksheet.dimensions

                )


            # =================================================
            # AUTO WIDTH
            # =================================================

            for column_cells in worksheet.columns:


                max_length = 0


                column_number = (

                    column_cells[0].column

                )


                column_letter = (

                    get_column_letter(
                        column_number
                    )

                )


                for cell in column_cells:


                    try:


                        cell_value = str(
                            cell.value
                        )


                        if (

                            len(cell_value)

                            > max_length

                        ):


                            max_length = (

                                len(
                                    cell_value
                                )

                            )


                    except Exception:

                        pass


                worksheet.column_dimensions[

                    column_letter

                ].width = min(

                    max_length + 2,

                    40

                )


    # ========================================================
    # FINAL SUMMARY
    # ========================================================

    print()

    print(
        "=" * 80
    )

    print(
        "SCRAPING SELESAI"
    )

    print(
        "=" * 80
    )


    print(

        f"Total Data Driver: "
        f"{len(df):,}"

    )


    if not df.empty:


        total_races = (

            df[

                [

                    "Season",

                    "Round"

                ]

            ]

            .drop_duplicates()

            .shape[0]

        )


        total_drivers = (

            df[

                "Driver Name"

            ]

            .nunique()

        )


        print(

            f"Total Race: "
            f"{total_races:,}"

        )


        print(

            f"Total Unique Driver: "
            f"{total_drivers:,}"

        )


    print(

        f"Unavailable Season: "
        f"{len(unavailable_df):,}"

    )


    print()


    print(
        f"File Output:"
    )


    print(
        OUTPUT_FILE
    )


    print(
        "=" * 80
    )


# ============================================================
# RUN
# ============================================================

if __name__ == "__main__":

    main()